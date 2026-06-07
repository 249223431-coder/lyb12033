import os
import uuid
import tempfile
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import db, CollectionForm, CollectionField, CollectionSubmission, CollectionValue
from config import Config

data_search_bp = Blueprint('data_search', __name__)


@data_search_bp.route('/')
@login_required
def index():
    forms = CollectionForm.query.filter_by(is_active=True).order_by(CollectionForm.sort_order).all()
    return render_template('data_search/index.html', forms=forms)


@data_search_bp.route('/form/<int:form_id>')
@login_required
def fill_form(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    if not form.is_active:
        flash('该采集表已停用', 'error')
        return redirect(url_for('data_search.index'))
    fields = form.fields.order_by(CollectionField.sort_order).all()
    return render_template('data_search/form.html', form=form, fields=fields)


@data_search_bp.route('/form/<int:form_id>/submit', methods=['POST'])
@login_required
def submit_form(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    if not form.is_active:
        flash('该采集表已停用', 'error')
        return redirect(url_for('data_search.index'))

    submission = CollectionSubmission(
        form_id=form_id,
        user_id=current_user.id
    )
    db.session.add(submission)
    db.session.flush()

    for field in form.fields.all():
        if field.field_type == 'file':
            file = request.files.get(f'field_{field.id}')
            if file and file.filename:
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                allowed = {'jpg', 'jpeg', 'png', 'gif', 'bmp', 'pdf', 'doc', 'docx', 'xls', 'xlsx'}
                if ext in allowed:
                    unique_name = f"collect_{uuid.uuid4().hex}.{ext}"
                    file.save(os.path.join(Config.UPLOAD_FOLDER, unique_name))
                    value = unique_name
                elif ext:
                    flash(f'字段「{field.label}」的文件格式不支持，已跳过', 'warning')
                    continue
                else:
                    continue
            else:
                if field.required:
                    flash(f'字段「{field.label}」为必填项', 'error')
                    db.session.rollback()
                    return redirect(url_for('data_search.fill_form', form_id=form_id))
                continue
        elif field.field_type == 'checkbox':
            vals = request.form.getlist(f'field_{field.id}')
            value = '\n'.join(vals) if vals else ''
            if field.required and not value:
                flash(f'字段「{field.label}」为必填项', 'error')
                db.session.rollback()
                return redirect(url_for('data_search.fill_form', form_id=form_id))
        else:
            value = request.form.get(f'field_{field.id}', '').strip()
            if field.required and not value:
                flash(f'字段「{field.label}」为必填项', 'error')
                db.session.rollback()
                return redirect(url_for('data_search.fill_form', form_id=form_id))

        cv = CollectionValue(
            submission_id=submission.id,
            field_id=field.id,
            value=value
        )
        db.session.add(cv)

    db.session.commit()
    flash('提交成功', 'success')
    return redirect(url_for('data_search.index'))


@data_search_bp.route('/my')
@login_required
def my_submissions():
    submissions = CollectionSubmission.query.filter_by(user_id=current_user.id)\
        .order_by(CollectionSubmission.created_at.desc()).all()
    return render_template('data_search/my_submissions.html', submissions=submissions)


@data_search_bp.route('/manage')
@login_required
def manage():
    forms = CollectionForm.query.order_by(CollectionForm.sort_order).all()
    return render_template('data_search/manage.html', forms=forms)


@data_search_bp.route('/manage/create', methods=['POST'])
@login_required
def create_form():
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    icon = request.form.get('icon', 'bi-clipboard').strip()
    if not title:
        flash('请输入表单标题', 'error')
        return redirect(url_for('data_search.manage'))
    form = CollectionForm(
        title=title,
        description=description,
        icon=icon,
        created_by=current_user.id,
        sort_order=CollectionForm.query.count()
    )
    db.session.add(form)
    db.session.commit()
    flash('采集表创建成功，请添加字段', 'success')
    return redirect(url_for('data_search.edit_form', form_id=form.id))


@data_search_bp.route('/manage/import', methods=['POST'])
@login_required
def import_form():
    """从Excel导入创建采集表"""
    title = request.form.get('title', '').strip()
    if not title:
        flash('请输入表单标题', 'error')
        return redirect(url_for('data_search.manage'))

    if 'excel_file' not in request.files:
        flash('请选择Excel文件', 'error')
        return redirect(url_for('data_search.manage'))

    file = request.files['excel_file']
    if not file or not file.filename:
        flash('请选择Excel文件', 'error')
        return redirect(url_for('data_search.manage'))

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ('xls', 'xlsx'):
        flash('仅支持 .xls / .xlsx 格式', 'error')
        return redirect(url_for('data_search.manage'))

    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file.read())
        ws = wb.sheet_by_index(0)
    except Exception as e:
        flash(f'Excel读取失败: {str(e)}', 'error')
        return redirect(url_for('data_search.manage'))

    if ws.nrows < 1:
        flash('Excel文件为空', 'error')
        return redirect(url_for('data_search.manage'))

    form = CollectionForm(
        title=title,
        description=f'从 {file.filename} 导入',
        icon='bi-file-earmark-excel',
        created_by=current_user.id,
        sort_order=CollectionForm.query.count()
    )
    db.session.add(form)
    db.session.flush()

    headers = []
    for c in range(min(ws.ncols, 50)):
        h = str(ws.cell_value(0, c)).strip()
        if h:
            headers.append(h)

    sample_row = None
    if ws.nrows > 1:
        sample_row = [str(ws.cell_value(1, c)).strip() for c in range(min(ws.ncols, 50))]

    for i, label in enumerate(headers):
        field_type = 'text'
        sample_val = sample_row[i] if sample_row and i < len(sample_row) else ''
        if sample_val:
            if sample_val.replace('.', '').replace('-', '').isdigit():
                if '.' in sample_val:
                    field_type = 'number'
                elif len(sample_val) == 8 and sample_val.isdigit():
                    field_type = 'date'
                else:
                    field_type = 'number'
            elif '-' in sample_val and len(sample_val) >= 8:
                field_type = 'date'
            elif len(sample_val) > 50:
                field_type = 'textarea'

        field = CollectionField(
            form_id=form.id,
            label=label,
            field_type=field_type,
            placeholder='',
            required=False,
            sort_order=i
        )
        db.session.add(field)

    db.session.commit()
    flash(f'导入成功！创建了 {len(headers)} 个字段，可在此页面微调', 'success')
    return redirect(url_for('data_search.edit_form', form_id=form.id))


@data_search_bp.route('/manage/template')
@login_required
def download_template():
    """下载Excel采集表模板"""
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl 未安装，请联系管理员', 'error')
        return redirect(url_for('data_search.manage'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '采集表'

    ws.append(['姓名', '部门', '联系电话', '备注'])
    ws.append(['张三', '生产部', '13800138000', '这是示例备注信息'])

    header_fill = openpyxl.styles.PatternFill(start_color='00897B', end_color='00897B', fill_type='solid')
    header_font = openpyxl.styles.Font(color='FFFFFF', bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    ws2 = wb.create_sheet('使用说明')
    ws2.append(['采集表Excel模板使用说明'])
    ws2.append(['1. 第一行是字段名称（列标题），系统会自动读取作为采集表的字段标签'])
    ws2.append(['2. 第二行可以填写示例数据，系统会自动识别字段类型：'])
    ws2.append(['   - 纯数字 → 数字输入框'])
    ws2.append(['   - 日期格式 → 日期选择器'])
    ws2.append(['   - 长文本（>50字符）→ 多行文本框'])
    ws2.append(['   - 短文本 → 单行文本框'])
    ws2.append(['3. 第二行可选，没有示例数据时所有字段默认为文本类型'])
    ws2.append(['4. 导入后可进入编辑页面微调字段类型和必填设置'])
    ws2.append(['5. 支持 .xlsx 和 .xls 格式'])
    ws2.column_dimensions['A'].width = 80

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True, download_name='采集表模板.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@data_search_bp.route('/manage/<int:form_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_form(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    if request.method == 'POST':
        form.title = request.form.get('title', '').strip()
        form.description = request.form.get('description', '').strip()
        form.icon = request.form.get('icon', 'bi-clipboard').strip()
        form.is_active = request.form.get('is_active') == 'on'
        db.session.commit()
        flash('采集表已更新', 'success')
        return redirect(url_for('data_search.manage'))
    fields = form.fields.order_by(CollectionField.sort_order).all()
    return render_template('data_search/edit_form.html', form=form, fields=fields)


@data_search_bp.route('/manage/<int:form_id>/delete', methods=['POST'])
@login_required
def delete_form(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    CollectionValue.query.filter(
        CollectionValue.submission_id.in_(
            db.session.query(CollectionSubmission.id).filter_by(form_id=form_id)
        )
    ).delete(synchronize_session='fetch')
    CollectionSubmission.query.filter_by(form_id=form_id).delete()
    CollectionField.query.filter_by(form_id=form_id).delete()
    db.session.delete(form)
    db.session.commit()
    flash('采集表已删除', 'success')
    return redirect(url_for('data_search.manage'))


@data_search_bp.route('/manage/<int:form_id>/field/add', methods=['POST'])
@login_required
def add_field(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    label = request.form.get('label', '').strip()
    field_type = request.form.get('field_type', 'text')
    if not label:
        flash('请输入字段名称', 'error')
        return redirect(url_for('data_search.edit_form', form_id=form_id))
    field = CollectionField(
        form_id=form_id,
        label=label,
        field_type=field_type,
        options=request.form.get('options', '').strip(),
        placeholder=request.form.get('placeholder', '').strip(),
        required=request.form.get('required') == 'on',
        sort_order=form.fields.count()
    )
    db.session.add(field)
    db.session.commit()
    flash(f'字段「{label}」已添加', 'success')
    return redirect(url_for('data_search.edit_form', form_id=form_id))


@data_search_bp.route('/manage/field/<int:field_id>/delete', methods=['POST'])
@login_required
def delete_field(field_id):
    field = CollectionField.query.get_or_404(field_id)
    form_id = field.form_id
    CollectionValue.query.filter_by(field_id=field_id).delete()
    db.session.delete(field)
    db.session.commit()
    flash('字段已删除', 'success')
    return redirect(url_for('data_search.edit_form', form_id=form_id))


@data_search_bp.route('/manage/<int:form_id>/submissions')
@login_required
def view_submissions(form_id):
    form = CollectionForm.query.get_or_404(form_id)
    submissions = form.submissions.order_by(CollectionSubmission.created_at.desc()).all()
    fields = form.fields.order_by(CollectionField.sort_order).all()

    data = []
    for sub in submissions:
        row = {'sub': sub}
        for field in fields:
            val = CollectionValue.query.filter_by(submission_id=sub.id, field_id=field.id).first()
            row[field.id] = val.value if val else ''
        data.append(row)

    return render_template('data_search/submissions.html', form=form, fields=fields, data=data)


@data_search_bp.route('/manage/submission/<int:submission_id>/delete', methods=['POST'])
@login_required
def delete_submission(submission_id):
    """所有成员都可以删除提交记录"""
    sub = CollectionSubmission.query.get_or_404(submission_id)
    form_id = sub.form_id
    CollectionValue.query.filter_by(submission_id=submission_id).delete()
    db.session.delete(sub)
    db.session.commit()
    flash('记录已删除', 'success')
    return redirect(url_for('data_search.view_submissions', form_id=form_id))
