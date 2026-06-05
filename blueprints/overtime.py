import io
import requests
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import db, OvertimePreReport
from datetime import datetime, date, timedelta

overtime_bp = Blueprint('overtime', __name__)


class HolidayCalendar:
    """中国法定节假日日历（对接timor.tech免费API）"""

    def __init__(self):
        self._cache = {}
        self._cache_year = None

    def refresh(self):
        try:
            year = date.today().year
            resp = requests.get(
                f'https://timor.tech/api/holiday/year/{year}',
                timeout=10,
                headers={'User-Agent': 'TeamManagement/1.0'}
            )
            data = resp.json()
            if data.get('code') == 0:
                self._cache = data['holiday']
                self._cache_year = year
        except Exception:
            pass

    def is_holiday(self, d):
        if self._cache_year != d.year:
            self.refresh()
        key = d.strftime('%m-%d')
        info = self._cache.get(key)
        if info and info.get('holiday'):
            return True, info.get('name', '法定节假日')
        return False, ''

    def is_weekend(self, d):
        return d.weekday() >= 5

    def get_upcoming_holidays(self, days=14):
        self.refresh()
        result = []
        today = date.today()
        for i in range(days):
            d = today + timedelta(days=i + 1)
            is_hol, name = self.is_holiday(d)
            is_weekend = self.is_weekend(d)
            if is_hol or is_weekend:
                weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
                result.append({
                    'date': d.strftime('%Y-%m-%d'),
                    'date_display': d.strftime('%m月%d日'),
                    'weekday': weekday_names[d.weekday()],
                    'type': 'holiday' if is_hol else 'weekend',
                    'name': name if is_hol else weekday_names[d.weekday()],
                })
        return result


holiday_calendar = HolidayCalendar()


@overtime_bp.route('/')
@login_required
def index():
    upcoming = holiday_calendar.get_upcoming_holidays()

    my_reports = OvertimePreReport.query.filter_by(
        user_id=current_user.id
    ).order_by(OvertimePreReport.holiday_date.asc()).all()

    reported_dates = {}
    for r in my_reports:
        if r.status == 'submitted':
            reported_dates[r.holiday_date.strftime('%Y-%m-%d')] = r

    return render_template('overtime/index.html',
                           upcoming=upcoming,
                           my_reports=my_reports,
                           reported_dates=reported_dates)


@overtime_bp.route('/submit', methods=['POST'])
@login_required
def submit():
    holiday_date = request.form.get('holiday_date', '')
    holiday_type = request.form.get('holiday_type', 'weekend')
    holiday_name = request.form.get('holiday_name', '')
    start_time = request.form.get('start_time', '')
    end_time = request.form.get('end_time', '')
    hours = request.form.get('hours', '0')
    reason = request.form.get('reason', '')

    if not holiday_date:
        flash('请选择加班日期', 'error')
        return redirect(url_for('overtime.index'))

    try:
        hours_val = float(hours) if hours else 0
    except ValueError:
        hours_val = 0

    if start_time and end_time and hours_val == 0:
        try:
            t1 = datetime.strptime(start_time, '%H:%M')
            t2 = datetime.strptime(end_time, '%H:%M')
            diff = t2 - t1
            hours_val = round(diff.total_seconds() / 3600, 1)
            if hours_val <= 0:
                hours_val = 0
        except Exception:
            hours_val = 0

    try:
        h_date = datetime.strptime(holiday_date, '%Y-%m-%d').date()
    except ValueError:
        flash('日期格式错误', 'error')
        return redirect(url_for('overtime.index'))

    existing = OvertimePreReport.query.filter_by(
        user_id=current_user.id,
        holiday_date=h_date,
        status='submitted'
    ).first()

    if existing:
        existing.holiday_type = holiday_type
        existing.holiday_name = holiday_name
        existing.start_time = start_time
        existing.end_time = end_time
        existing.hours = hours_val
        existing.reason = reason
        flash('加班预提报已更新', 'success')
    else:
        record = OvertimePreReport(
            user_id=current_user.id,
            holiday_date=h_date,
            holiday_type=holiday_type,
            holiday_name=holiday_name,
            start_time=start_time,
            end_time=end_time,
            hours=hours_val,
            reason=reason,
            status='submitted'
        )
        db.session.add(record)
        flash('加班预提报提交成功', 'success')

    db.session.commit()
    return redirect(url_for('overtime.index'))


@overtime_bp.route('/delete/<int:report_id>', methods=['POST'])
@login_required
def delete(report_id):
    report = OvertimePreReport.query.get_or_404(report_id)
    if report.user_id != current_user.id and not current_user.is_admin:
        flash('无权操作', 'error')
        return redirect(url_for('overtime.index'))
    report.status = 'cancelled'
    db.session.commit()
    flash('加班预提报已取消', 'success')
    return redirect(url_for('overtime.index'))


@overtime_bp.route('/summaries')
@login_required
def summaries():
    """显示所有已提报记录，按日期分组（所有成员可见，管理员可导出）"""
    reports = OvertimePreReport.query.filter_by(status='submitted') \
        .order_by(OvertimePreReport.holiday_date.asc(),
                  OvertimePreReport.user_id.asc()).all()

    grouped = {}
    for r in reports:
        key = r.holiday_date.strftime('%Y-%m-%d')
        if key not in grouped:
            grouped[key] = {
                'date': r.holiday_date,
                'date_display': r.holiday_date.strftime('%Y-%m-%d'),
                'holiday_type': r.holiday_type,
                'holiday_name': r.holiday_name,
                'reports': [],
                'total_hours': 0,
            }
        grouped[key]['reports'].append(r)
        grouped[key]['total_hours'] += r.hours or 0

    return render_template('overtime/summaries.html', grouped=grouped)


@overtime_bp.route('/export_all')
@login_required
def export_all():
    """管理员导出所有 submitted 记录并清空"""
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('overtime.summaries'))

    reports = OvertimePreReport.query.filter_by(status='submitted') \
        .order_by(OvertimePreReport.holiday_date.asc(),
                  OvertimePreReport.user_id.asc()).all()

    if not reports:
        flash('没有可导出的提报记录', 'error')
        return redirect(url_for('overtime.summaries'))

    # 生成 Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = '加班预提汇总'

    title_text = f'加班预提汇总表（{datetime.now().strftime("%Y-%m-%d")}导出）'
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    columns = ['序号', '工号', '姓名', '加班日期', '节假日类型', '时间', '小时数', '加班事由']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 3
    current_date = None

    for r in reports:
        r_date = r.holiday_date.strftime('%Y-%m-%d') if r.holiday_date else ''
        # 每个日期组之间插入分组标题行
        if r_date != current_date:
            current_date = r_date
            group_title = f'{r_date}  {r.holiday_name} 加班预提'
            ws.merge_cells(f'A{row}:H{row}')
            group_cell = ws.cell(row=row, column=1, value=group_title)
            group_cell.font = Font(bold=True, size=11, color='2B579A')
            group_cell.fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
            group_cell.alignment = Alignment(horizontal='left', vertical='center')
            for c in range(1, 9):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

            # 表头
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            row += 1

        time_str = f'{r.start_time or ""}-{r.end_time or ""}' if r.start_time else '-'
        values = [
            row - 2,
            r.user.username if r.user else '',
            r.user.real_name if r.user else '',
            r_date,
            r.holiday_name or ('周末' if r.holiday_type == 'weekend' else '法定节假日'),
            time_str,
            r.hours or '',
            r.reason or '',
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        row += 1

    # 列宽
    col_widths = [6, 12, 12, 14, 14, 16, 10, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 将所有记录标记为 exported
    for r in reports:
        r.status = 'exported'
    db.session.commit()

    count = len(reports)
    download_name = f'加班预提汇总_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    flash(f'已导出 {count} 条加班预提记录，记录已清空', 'success')
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=download_name)
