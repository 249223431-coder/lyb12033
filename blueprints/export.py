import io
from flask import Blueprint, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from models import AttendanceRecord, InspectionRecord, AnomalyReport, SparePart, StockTransaction, Document, PurchaseRequisition
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

export_bp = Blueprint('export', __name__)


def _create_excel(columns, data, sheet_name='数据'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col_idx in range(1, len(columns) + 1):
        max_length = len(str(columns[col_idx - 1]))
        for row_data in data:
            val = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ''
            max_length = max(max_length, len(val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@export_bp.route('/attendance')
@login_required
def export_attendance():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('attendance.index'))

    records = AttendanceRecord.query.order_by(AttendanceRecord.created_at.desc()).all()
    columns = ['姓名', '类型', '开始日期', '结束日期', '开始时间', '结束时间', '天数', '原因', '状态', '记录时间']
    data = []
    for r in records:
        data.append([
            r.user.real_name if r.user else '', '请假' if r.record_type == 'leave' else '加班',
            r.start_date.strftime('%Y-%m-%d') if r.start_date else '',
            r.end_date.strftime('%Y-%m-%d') if r.end_date else '',
            r.start_time, r.end_time, r.duration_days, r.reason, r.status,
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        ])

    output = _create_excel(columns, data, '考勤记录')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'考勤记录_{datetime.now().strftime("%Y%m%d")}.xlsx')


@export_bp.route('/inspection')
@login_required
def export_inspection():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('inspection.records'))

    records = InspectionRecord.query.order_by(InspectionRecord.created_at.desc()).limit(5000).all()
    columns = ['点检项目', '位置', '检查人', '检查日期', '检查值', '正常范围', '是否正常', '备注', '记录时间']
    data = []
    for r in records:
        data.append([
            r.item.name if r.item else '', r.item.location if r.item else '',
            r.user.real_name if r.user else '',
            r.check_date.strftime('%Y-%m-%d') if r.check_date else '',
            r.check_value, r.item.normal_range if r.item else '',
            '正常' if r.is_normal else '异常', r.remark,
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        ])

    output = _create_excel(columns, data, '点检记录')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'点检记录_{datetime.now().strftime("%Y%m%d")}.xlsx')


@export_bp.route('/anomaly')
@login_required
def export_anomaly():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('inspection.anomaly_list'))

    reports = AnomalyReport.query.order_by(AnomalyReport.created_at.desc()).all()
    columns = ['等级', '标题', '位置', '描述', '提报人', '状态', '处理方式', '处理备注', '提报时间']
    data = []
    for r in reports:
        action_map = {'immediate': '立即处理', 'observe': '待观察', 'shutdown': '停机处理'}
        data.append([
            '紧急' if r.level == 'urgent' else '一般',
            r.title, r.location, r.description,
            r.user.real_name if r.user else '',
            {'open': '待处理', 'processing': '处理中', 'resolved': '已解决'}.get(r.status, r.status),
            action_map.get(r.handle_action, ''),
            r.handler_note,
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        ])

    output = _create_excel(columns, data, '异常报告')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'异常报告_{datetime.now().strftime("%Y%m%d")}.xlsx')


@export_bp.route('/spare_parts_stock')
@login_required
def export_spare_parts_stock():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('spare_parts.index'))

    parts = SparePart.query.filter_by(is_active=True).order_by(SparePart.name).all()
    columns = ['编码', '名称', '型号规格', '单位', '类别', '库位', '库存数量', '安全库存', '预警库存', '单价', '供应商', '备注']
    data = []
    for p in parts:
        data.append([p.part_code, p.name, p.model_spec, p.unit, p.category, p.location,
                     p.stock_quantity, p.safety_stock, p.warning_stock, p.price, p.supplier, p.remark])

    output = _create_excel(columns, data, '备件库存')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'备件库存_{datetime.now().strftime("%Y%m%d")}.xlsx')


@export_bp.route('/spare_parts_transactions')
@login_required
def export_spare_parts_transactions():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('spare_parts.transactions'))

    records = StockTransaction.query.order_by(StockTransaction.created_at.desc()).limit(5000).all()
    columns = ['备件编码', '备件名称', '类型', '数量', '操作人', '用途', '备注', '操作时间']
    data = []
    for r in records:
        data.append([
            r.part.part_code if r.part else '', r.part.name if r.part else '',
            '入库' if r.trans_type == 'in' else '出库', r.quantity,
            r.operator_name, r.purpose, r.remark,
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        ])

    output = _create_excel(columns, data, '出入库记录')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'出入库记录_{datetime.now().strftime("%Y%m%d")}.xlsx')


@export_bp.route('/purchase')
@login_required
def export_purchase():
    if not current_user.is_admin:
        flash('需要管理员权限', 'error')
        return redirect(url_for('spare_parts.purchase_list'))

    records = PurchaseRequisition.query.order_by(PurchaseRequisition.created_at.desc()).all()
    columns = ['编码', '备件名称', '型号规格', '数量', '请购目的', '提交人', '状态', '提交时间']
    data = []
    for r in records:
        data.append([
            r.part_code or '', r.part_name, r.model_spec, r.quantity, r.purpose,
            r.user.real_name if r.user else '',
            {'pending': '待确认', 'confirmed': '已确认', 'cancelled': '已取消'}.get(r.status, r.status),
            r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''
        ])

    output = _create_excel(columns, data, '请购记录')
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'请购记录_{datetime.now().strftime("%Y%m%d")}.xlsx')
