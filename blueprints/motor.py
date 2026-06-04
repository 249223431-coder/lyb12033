import os, uuid
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import db, Motor, MotorRepair
from config import Config
from blueprints.auth import admin_required
import openpyxl
import io
from datetime import datetime
from copy import copy

motor_bp = Blueprint('motor', __name__)


@motor_bp.route('/')
@login_required
def index():
    return render_template('motor/index.html')


@motor_bp.route('/search')
@login_required
def search():
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return render_template('motor/search.html', motors=[], query='')
    motors = Motor.query.filter(
        Motor.name.contains(q) | Motor.motor_code.contains(q) |
        Motor.motor_model.contains(q) | Motor.location.contains(q) |
        Motor.room.contains(q) | Motor.cabinet.contains(q) |
        Motor.position.contains(q)
    ).order_by(Motor.motor_code).limit(50).all()
    return render_template('motor/search.html', motors=motors, query=q)


@motor_bp.route('/repair', methods=['GET', 'POST'])
@login_required
def repair():
    if request.method == 'POST':
        data = {}
        fields = [
            'repair_name', 'use_location', 'pole_count', 'power_spec', 'model_spec',
            'fault_desc', 'sap_price', 'repair_quote', 'vendor',
            'repair_number', 'aplus_notice', 'aplus_budget',
            'cost_center', 'opex_no', 'opex_line', 'acceptance_result',
            'approver_gm', 'approver_director', 'approver_manager',
            'approver_chief', 'approver_section_chief', 'approver_handler',
            'acceptor_chief', 'acceptor_handler',
        ]
        for f in fields:
            data[f] = request.form.get(f, '').strip()

        if not data['repair_name']:
            flash('请填写委外维修件名称', 'error')
            return redirect(url_for('motor.repair'))

        repair = MotorRepair(user_id=current_user.id, **data)
        db.session.add(repair)
        db.session.commit()
        flash('维修申请已保存', 'success')
        return redirect(url_for('motor.repair_list'))
    return render_template('motor/repair_form.html')


@motor_bp.route('/repair/list')
@login_required
def repair_list():
    records = MotorRepair.query.order_by(MotorRepair.created_at.desc()).all()
    return render_template('motor/repair_list.html', records=records)


@motor_bp.route('/repair/<int:rid>/export')
@login_required
def repair_export(rid):
    record = MotorRepair.query.get_or_404(rid)
    template_path = r'c:\Users\lyb_app\Desktop\电机委外维修.xlsx'
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    mapping = {
        (4, 3): record.repair_name,
        (4, 7): record.use_location,
        (5, 3): record.pole_count,
        (5, 5): record.power_spec,
        (5, 7): record.model_spec,
        (6, 3): record.fault_desc,
        (8, 3): record.sap_price,
        (8, 5): record.repair_quote,
        (8, 7): record.vendor,
        (10, 3): record.approver_gm,
        (10, 5): record.approver_director,
        (10, 8): record.approver_manager,
        (11, 3): record.approver_chief,
        (11, 5): record.approver_section_chief,
        (11, 8): record.approver_handler,
        (13, 3): record.acceptance_result,
        (13, 6): record.acceptor_chief,
        (13, 8): record.acceptor_handler,
        (14, 3): record.repair_number,
        (14, 6): record.aplus_notice,
        (14, 8): record.aplus_budget,
        (15, 3): record.cost_center,
        (15, 6): record.opex_no,
        (15, 8): record.opex_line,
    }
    for (r, c), val in mapping.items():
        if val:
            ws.cell(row=r, column=c, value=val)

    today_str = datetime.now().strftime('%Y年%m月%d日')
    ws.cell(row=3, column=6, value=today_str)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'电机委外维修_{record.repair_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@motor_bp.route('/repair/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def repair_edit(rid):
    record = MotorRepair.query.get_or_404(rid)
    if request.method == 'POST':
        fields = [
            'repair_name', 'use_location', 'pole_count', 'power_spec', 'model_spec',
            'fault_desc', 'sap_price', 'repair_quote', 'vendor',
            'repair_number', 'aplus_notice', 'aplus_budget',
            'cost_center', 'opex_no', 'opex_line', 'acceptance_result',
            'approver_gm', 'approver_director', 'approver_manager',
            'approver_chief', 'approver_section_chief', 'approver_handler',
            'acceptor_chief', 'acceptor_handler',
        ]
        for f in fields:
            setattr(record, f, request.form.get(f, '').strip())
        db.session.commit()
        flash('维修申请已更新', 'success')
        return redirect(url_for('motor.repair_list'))
    return render_template('motor/repair_form.html', record=record)


@motor_bp.route('/repair/<int:rid>/delete', methods=['POST'])
@login_required
@admin_required
def repair_delete(rid):
    record = MotorRepair.query.get_or_404(rid)
    db.session.delete(record)
    db.session.commit()
    flash('记录已删除', 'success')
    return redirect(url_for('motor.repair_list'))
