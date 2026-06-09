import os
import re
import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from models import db, ShutdownPlan, ShutdownPlanItem, Notification, User
from datetime import datetime, date
from blueprints.auth import admin_required

shutdown_bp = Blueprint('shutdown', __name__)

EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), '停机计划.xlsx')


def find_user_by_name(name_text):
    """根据姓名匹配系统用户"""
    name_text = name_text.strip()
    if not name_text:
        return None
    user = User.query.filter_by(real_name=name_text, is_active=True).first()
    if not user:
        user = User.query.filter(
            User.real_name.contains(name_text),
            User.is_active == True
        ).first()
    return user


def parse_member_names(members_text):
    """解析成员列，按空格拆分取第一个词作为人名"""
    if not members_text:
        return []
    names = []
    for part in members_text.strip().split():
        if len(part) >= 2 and not part.startswith('PM') and part not in ['完成', '电气', '龙鼎电气']:
            names.append(part)
    return names


def get_latest_plan():
    """获取最新的停机计划"""
    return ShutdownPlan.query.order_by(ShutdownPlan.created_at.desc()).first()


def load_excel_sheet():
    """读取Excel最后一个非空Sheet的数据"""
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH, data_only=True)
    # 跳过空Sheet1，取倒数第二个
    sheet_names = [s for s in wb.sheetnames if s != 'Sheet1']
    if not sheet_names:
        wb.close()
        return None, None, None
    sheet_name = sheet_names[-1]
    ws = wb[sheet_name]

    # 从第2行读取元信息：日期和时间
    meta_text = str(ws.cell(row=2, column=1).value or '')
    plan_date_str = None
    start_time = '08:00'
    end_time = '18:00'
    date_match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', meta_text)
    if date_match:
        plan_date_str = f'{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}'
    time_match = re.search(r'(\d{1,2}:\d{2})[—\-－](\d{1,2}:\d{2})', meta_text)
    if time_match:
        start_time = time_match.group(1)
        end_time = time_match.group(2)

    department = str(ws.cell(row=3, column=1).value or '').replace('提出部门：', '').strip()

    # 从第5行开始读取数据（第4行是表头）
    items_data = []
    seq_counter = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        seq = row[0] if len(row) > 0 else None
        if seq is None or str(seq).strip() == '':
            continue
        try:
            seq_counter = int(seq)
        except (ValueError, TypeError):
            continue

        project_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        category = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        responsible = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        members = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        safety_notes = str(row[5]).strip() if len(row) > 5 and row[5] else ''

        items_data.append({
            'seq': seq_counter,
            'project_name': project_name,
            'category': category,
            'responsible': responsible,
            'members': members,
            'safety_notes': safety_notes,
        })

    wb.close()
    return sheet_name, items_data, {
        'plan_date_str': plan_date_str,
        'start_time': start_time,
        'end_time': end_time,
        'department': department,
    }


@shutdown_bp.route('/')
@login_required
def index():
    plan = get_latest_plan()
    if plan:
        # 按负责人分组
        items = plan.items.all()
        grouped = {}
        for item in items:
            key = item.responsible or '未分配'
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(item)
        total = len(items)
        completed = sum(1 for it in items if it.completed)
    else:
        grouped = {}
        total = 0
        completed = 0

    return render_template('shutdown/index.html',
                           plan=plan, grouped=grouped,
                           total=total, completed=completed)


@shutdown_bp.route('/load', methods=['POST'])
@login_required
@admin_required
def load():
    plan_date_str = request.form.get('plan_date', '')
    start_time = request.form.get('start_time', '')
    end_time = request.form.get('end_time', '')

    if not plan_date_str:
        flash('请选择计划停机日期', 'error')
        return redirect(url_for('shutdown.index'))

    try:
        plan_date = datetime.strptime(plan_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('日期格式错误', 'error')
        return redirect(url_for('shutdown.index'))

    sheet_name, items_data, meta = load_excel_sheet()
    if not items_data:
        flash('无法读取Excel停机计划，请检查文件', 'error')
        return redirect(url_for('shutdown.index'))

    # 使用前端传入的日期和时间覆盖Excel中的
    final_start = start_time or meta['start_time']
    final_end = end_time or meta['end_time']

    # 删除已有计划
    old = ShutdownPlan.query.first()
    if old:
        ShutdownPlanItem.query.filter_by(plan_id=old.id).delete()
        db.session.delete(old)
        db.session.flush()

    # 创建计划
    plan = ShutdownPlan(
        sheet_name=sheet_name,
        plan_date=plan_date,
        start_time=final_start,
        end_time=final_end,
        department=meta['department'],
        status='loaded'
    )
    db.session.add(plan)
    db.session.flush()

    for d in items_data:
        # 匹配负责人
        resp_user = find_user_by_name(d['responsible'])
        item = ShutdownPlanItem(
            plan_id=plan.id,
            seq=d['seq'],
            project_name=d['project_name'],
            category=d['category'],
            responsible=d['responsible'],
            responsible_user_id=resp_user.id if resp_user else None,
            members=d['members'],
            safety_notes=d['safety_notes'],
        )
        db.session.add(item)

    db.session.commit()
    flash(f'已加载停机计划 [{sheet_name}]，共 {len(items_data)} 条项目', 'success')
    return redirect(url_for('shutdown.index'))


@shutdown_bp.route('/push', methods=['POST'])
@login_required
@admin_required
def push():
    plan = get_latest_plan()
    if not plan:
        flash('没有可推送的计划', 'error')
        return redirect(url_for('shutdown.index'))

    items = plan.items.all()
    # 按用户收集任务
    user_tasks = {}  # {user_id: [items...]}
    for item in items:
        if item.responsible_user_id:
            uid = item.responsible_user_id
            if uid not in user_tasks:
                user_tasks[uid] = []
            user_tasks[uid].append(item)

        for member_name in parse_member_names(item.members):
            member = find_user_by_name(member_name)
            if member and member.id not in user_tasks:
                user_tasks[member.id] = []
            if member:
                user_tasks[member.id].append(item)

    count = 0
    for uid, task_items in user_tasks.items():
        user = User.query.get(uid)
        if not user:
            continue
        # 去重
        seen = set()
        unique_items = []
        for it in task_items:
            if it.id not in seen:
                seen.add(it.id)
                unique_items.append(it)

        lines = []
        for it in unique_items:
            name = it.project_name or '(无名称)'
            safety = f'  [注意：{it.safety_notes}]' if it.safety_notes else ''
            lines.append(f'  · {name}{safety}')

        # 生成简洁的文本格式内容
        task_lines = []
        for idx, it in enumerate(unique_items, 1):
            name = it.project_name or '(无名称)'
            safety = it.safety_notes
            line = f'{idx}. {name}'
            if safety:
                line += f' 【注意：{safety}】'
            task_lines.append(line)
        
        content = f'停机日期：{plan.plan_date.strftime("%Y-%m-%d")} {plan.start_time}-{plan.end_time}\n\n任务清单：\n' + '\n'.join(task_lines)
        notification = Notification(
            user_id=uid,
            title=f'停机计划任务 - {plan.plan_date.strftime("%Y-%m-%d")}',
            content=content
        )
        db.session.add(notification)
        count += 1

    plan.status = 'notified'
    db.session.commit()
    flash(f'已向 {count} 位成员推送停机任务通知', 'success')
    return redirect(url_for('shutdown.index'))


@shutdown_bp.route('/edit', methods=['POST'])
@login_required
@admin_required
def edit():
    plan = get_latest_plan()
    if not plan:
        flash('没有可编辑的计划', 'error')
        return redirect(url_for('shutdown.index'))

    plan_date_str = request.form.get('plan_date', '')
    start_time = request.form.get('start_time', '')
    end_time = request.form.get('end_time', '')

    if plan_date_str:
        try:
            plan.plan_date = datetime.strptime(plan_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if start_time:
        plan.start_time = start_time
    if end_time:
        plan.end_time = end_time

    db.session.commit()
    flash('停机计划信息已更新', 'success')
    return redirect(url_for('shutdown.index'))


@shutdown_bp.route('/confirm/<int:item_id>', methods=['POST'])
@login_required
def confirm(item_id):
    item = ShutdownPlanItem.query.get_or_404(item_id)
    note = request.form.get('complete_note', '')
    item.completed = True
    item.complete_note = note
    item.confirmed_by = current_user.id
    item.confirmed_at = datetime.utcnow()
    db.session.commit()
    flash('项目已确认完成', 'success')
    return redirect(request.referrer or url_for('shutdown.index'))


@shutdown_bp.route('/cancel_confirm/<int:item_id>', methods=['POST'])
@login_required
def cancel_confirm(item_id):
    item = ShutdownPlanItem.query.get_or_404(item_id)
    item.completed = False
    item.complete_note = ''
    item.confirmed_by = None
    item.confirmed_at = None
    db.session.commit()
    flash('已撤销完成确认', 'success')
    return redirect(request.referrer or url_for('shutdown.index'))


@shutdown_bp.route('/my_tasks')
@login_required
def my_tasks():
    plan = get_latest_plan()
    if not plan:
        return render_template('shutdown/my_tasks.html', plan=None, items=[])

    # 查找当前用户相关的所有项目
    items = plan.items.all()
    my_items = []
    for item in items:
        is_mine = False
        role = ''
        if item.responsible_user_id == current_user.id:
            is_mine = True
            role = '负责人'
        else:
            for member_name in parse_member_names(item.members):
                member = find_user_by_name(member_name)
                if member and member.id == current_user.id:
                    is_mine = True
                    role = '成员'
                    break
        if is_mine:
            my_items.append({'item': item, 'role': role})

    return render_template('shutdown/my_tasks.html', plan=plan, my_items=my_items)


@shutdown_bp.route('/search')
@login_required
def search():
    keyword = request.args.get('q', '').strip()
    responsible = request.args.get('responsible', '')
    status_filter = request.args.get('status', '')

    query = ShutdownPlanItem.query.join(ShutdownPlan)

    if keyword:
        like = f'%{keyword}%'
        query = query.filter(
            db.or_(
                ShutdownPlanItem.project_name.ilike(like),
                ShutdownPlanItem.safety_notes.ilike(like),
                ShutdownPlanItem.responsible.ilike(like),
                ShutdownPlanItem.members.ilike(like),
            )
        )

    if responsible:
        query = query.filter(ShutdownPlanItem.responsible.ilike(f'%{responsible}%'))

    if status_filter == 'completed':
        query = query.filter(ShutdownPlanItem.completed == True)
    elif status_filter == 'pending':
        query = query.filter(ShutdownPlanItem.completed == False)

    results = query.order_by(ShutdownPlan.plan_date.desc()).limit(100).all()

    # 收集所有负责人供筛选
    all_responsibles = db.session.query(ShutdownPlanItem.responsible).distinct().all()
    responsibles = sorted(set(r[0] for r in all_responsibles if r[0]))

    return render_template('shutdown/search.html',
                           results=results,
                           keyword=keyword,
                           responsibles=responsibles,
                           responsible=responsible,
                           status_filter=status_filter)


@shutdown_bp.route('/summary')
@login_required
@admin_required
def summary():
    plan = get_latest_plan()
    if not plan:
        flash('没有停机计划', 'error')
        return redirect(url_for('shutdown.index'))
    items = plan.items.order_by(ShutdownPlanItem.seq).all()
    total = len(items)
    completed = sum(1 for it in items if it.completed)
    return render_template('shutdown/summary.html',
                           plan=plan, items=items,
                           total=total, completed=completed)


@shutdown_bp.route('/export')
@login_required
@admin_required
def export():
    plan = get_latest_plan()
    if not plan:
        flash('没有可导出的计划', 'error')
        return redirect(url_for('shutdown.index'))

    items = plan.items.order_by(ShutdownPlanItem.seq).all()
    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_PATH)
    sheet_name = plan.sheet_name
    if sheet_name not in wb.sheetnames:
        flash(f'Sheet [{sheet_name}] 不存在于Excel中', 'error')
        return redirect(url_for('shutdown.summary'))

    ws = wb[sheet_name]

    # 写回H列：完成确认
    col_h = 8
    for item in items:
        row = item.seq + 4  # 第4行是表头，数据从第5行开始
        if row <= ws.max_row:
            if item.completed:
                confirmer_name = ''
                if item.confirmed_by:
                    u = User.query.get(item.confirmed_by)
                    if u:
                        confirmer_name = u.real_name
                text = f'已完成'
                if item.complete_note:
                    text += f'（{item.complete_note}）'
                if confirmer_name:
                    text += f' 确认人:{confirmer_name}'
                ws.cell(row=row, column=col_h, value=text)
            else:
                ws.cell(row=row, column=col_h, value='未完成')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    plan.status = 'completed'
    db.session.commit()

    download_name = f'停机计划_{plan.sheet_name}_完成情况.xlsx'
    flash(f'已导出 {plan.sheet_name} 停机计划完成情况', 'success')
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=download_name)
